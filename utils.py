import csv
import pprint
import datetime
from openpyxl import load_workbook


class ParseNQuery:
    # parser_format_dict = {
    #     'xls': search_xls,
    #     'xlsx': search_xls,
    #     'xlsm': search_xls
    # }

    def __init__(self, name):
        self.name = name
        self.doc = load_workbook(filename=self.name).active
        self.headers = None

        ws = self.doc
        header_row = None
        for r in ws.iter_rows(min_row=1):
            if r[0].value:
                header_row = r
                break
        self.headers = [c.value for c in header_row]

    def parse_row_as_dict(self, row):
        obj = {}
        for c in range(len(self.headers)):
            if isinstance(row[c].value, datetime.datetime):
                dt: datetime.datetime = row[c].value
                obj[self.headers[c]] = dt.strftime("%d-%m-%y::%H:%M")
            else:
                obj[self.headers[c]] = row[c].value
        return obj

    def search_rows(self, row, term):
        for col in row:
            if term in str(col.value).lower():
                return True
        return False

    def search_xls(self, term):
        res = []
        ws = self.doc
        for row in ws.iter_rows():
            if self.search_rows(row, term.lower()):
                res.append(self.parse_row_as_dict(row))
        return res


# obj = ParseNQuery(name='data.xlsx')
# pprint.pp(obj.search_xls('frontend'))
